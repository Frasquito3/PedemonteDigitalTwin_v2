from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, time, timedelta
from math import ceil, isfinite
from pathlib import Path
import re
import unicodedata
from typing import Any, Iterable

from openpyxl import load_workbook


VERSION_IMPORTADOR = "pedemonte-orders-xlsx-v1"
HOJA_PEDIDOS_PREFERIDA = "Pedidos"

HORARIOS_TURNO: dict[str, tuple[int, int]] = {
    "MANANA": (7 * 60 + 30, 12 * 60),
    "TARDE": (14 * 60, 17 * 60),
}


@dataclass(frozen=True)
class ErrorFilaExcel:
    fila: int
    campo: str
    mensaje: str


@dataclass(frozen=True)
class PedidoExcel:
    fila_excel: int
    pedido_id: str
    cliente: str
    direccion: str
    barrio: str
    unidades: int
    latitud: float
    longitud: float
    requiere_volcador: bool
    tiene_ventana: bool
    hora_desde_min: int
    hora_hasta_min: int
    observaciones: str


@dataclass(frozen=True)
class ResultadoImportacionExcel:
    archivo: Path
    hoja: str
    turno: str
    pedidos: tuple[PedidoExcel, ...]
    tareas_estimadas: int
    filas_ignoradas: int
    version: str = VERSION_IMPORTADOR


class ErrorImportacionExcel(ValueError):
    def __init__(
        self,
        errores: Iterable[ErrorFilaExcel],
    ) -> None:
        self.errores = tuple(errores)

        if not self.errores:
            mensaje = "La planilla no pudo validarse."
        else:
            mensaje = " | ".join(
                (
                    f"fila {error.fila}, "
                    f"{error.campo}: {error.mensaje}"
                )
                for error in self.errores
            )

        super().__init__(mensaje)


_ALIAS_COLUMNAS: dict[str, set[str]] = {
    "pedido_id": {
        "pedidoid",
        "idpedido",
        "id",
        "codigo",
    },
    "cliente": {
        "cliente",
        "nombrecliente",
    },
    "direccion": {
        "direccion",
        "domicilio",
    },
    "barrio": {
        "barrio",
        "zona",
    },
    "unidades": {
        "unidades",
        "unidadescapacidad",
        "cantidad",
        "carga",
    },
    "latitud": {
        "latitud",
        "lat",
    },
    "longitud": {
        "longitud",
        "lon",
        "lng",
    },
    "requiere_volcador": {
        "requierevolcador",
        "volcador",
        "requierevolquete",
    },
    "tiene_ventana": {
        "tieneventana",
        "ventana",
        "tieneventanaespecifica",
    },
    "hora_desde": {
        "horadesde",
        "desde",
        "inicioventana",
    },
    "hora_hasta": {
        "horahasta",
        "hasta",
        "finventana",
    },
    "observaciones": {
        "observaciones",
        "observacion",
        "notas",
    },
}

_COLUMNAS_OBLIGATORIAS = {
    "pedido_id",
    "unidades",
    "latitud",
    "longitud",
    "requiere_volcador",
    "tiene_ventana",
    "hora_desde",
    "hora_hasta",
}


_RE_HORA = re.compile(
    r"^(?P<hora>[01]?\d|2[0-3]):(?P<minuto>[0-5]\d)$"
)


def importar_pedidos_excel(
    archivo: str | Path,
    *,
    turno: str,
    capacidad_camion: int = 8,
    max_tareas: int = 30,
    hoja: str | None = None,
) -> ResultadoImportacionExcel:
    ruta = Path(archivo).expanduser().resolve()

    if not ruta.is_file():
        raise FileNotFoundError(
            f"No existe la planilla: {ruta}"
        )

    if ruta.suffix.lower() != ".xlsx":
        raise ValueError(
            "Solo se admiten archivos .xlsx."
        )

    turno_normalizado = _normalizar_turno(turno)

    if capacidad_camion <= 0:
        raise ValueError(
            "capacidad_camion debe ser > 0."
        )

    if max_tareas <= 0:
        raise ValueError(
            "max_tareas debe ser > 0."
        )

    libro = load_workbook(
        ruta,
        read_only=True,
        data_only=True,
    )

    try:
        hoja_excel = _seleccionar_hoja(
            libro,
            hoja,
        )

        fila_encabezado, columnas = (
            _detectar_encabezado(
                hoja_excel.iter_rows(
                    min_row=1,
                    max_row=20,
                    values_only=True,
                )
            )
        )

        faltantes = sorted(
            _COLUMNAS_OBLIGATORIAS
            - set(columnas)
        )

        if faltantes:
            raise ErrorImportacionExcel(
                [
                    ErrorFilaExcel(
                        fila=fila_encabezado,
                        campo="encabezado",
                        mensaje=(
                            "faltan columnas obligatorias: "
                            + ", ".join(faltantes)
                        ),
                    )
                ]
            )

        pedidos: list[PedidoExcel] = []
        errores: list[ErrorFilaExcel] = []
        filas_ignoradas = 0
        ids_vistos: dict[str, int] = {}

        for numero_fila, valores in enumerate(
            hoja_excel.iter_rows(
                min_row=fila_encabezado + 1,
                values_only=True,
            ),
            start=fila_encabezado + 1,
        ):
            fila = {
                nombre: (
                    valores[indice]
                    if indice < len(valores)
                    else None
                )
                for nombre, indice in columnas.items()
            }

            if _fila_vacia(fila.values()):
                filas_ignoradas += 1
                continue

            pedido_id_crudo = _texto(
                fila.get("pedido_id")
            )
            fila_anterior = (
                ids_vistos.get(pedido_id_crudo)
                if pedido_id_crudo
                else None
            )
            duplicado = fila_anterior is not None

            if pedido_id_crudo and not duplicado:
                ids_vistos[pedido_id_crudo] = numero_fila

            if duplicado:
                errores.append(
                    ErrorFilaExcel(
                        fila=numero_fila,
                        campo="pedido_id",
                        mensaje=(
                            "ID duplicado; ya aparece en "
                            f"la fila {fila_anterior}."
                        ),
                    )
                )

            pedido, errores_fila = _parsear_fila(
                numero_fila,
                fila,
                turno=turno_normalizado,
            )

            errores.extend(errores_fila)

            if pedido is None or duplicado:
                continue

            pedidos.append(pedido)

        if not pedidos and not errores:
            errores.append(
                ErrorFilaExcel(
                    fila=fila_encabezado + 1,
                    campo="pedidos",
                    mensaje=(
                        "no se encontraron filas de pedidos."
                    ),
                )
            )

        tareas_estimadas = sum(
            ceil(
                pedido.unidades
                / capacidad_camion
            )
            for pedido in pedidos
        )

        if tareas_estimadas > max_tareas:
            errores.append(
                ErrorFilaExcel(
                    fila=fila_encabezado,
                    campo="pedidos",
                    mensaje=(
                        "la importación generaría "
                        f"{tareas_estimadas} tareas después "
                        "del split y supera el máximo "
                        f"permitido de {max_tareas}."
                    ),
                )
            )

        if errores:
            raise ErrorImportacionExcel(errores)

        return ResultadoImportacionExcel(
            archivo=ruta,
            hoja=hoja_excel.title,
            turno=turno_normalizado,
            pedidos=tuple(pedidos),
            tareas_estimadas=tareas_estimadas,
            filas_ignoradas=filas_ignoradas,
        )

    finally:
        libro.close()


def parsear_hora_a_minutos(valor: Any) -> int:
    if isinstance(valor, datetime):
        return valor.hour * 60 + valor.minute

    if isinstance(valor, time):
        return valor.hour * 60 + valor.minute

    if isinstance(valor, timedelta):
        minutos = valor.total_seconds() / 60.0
        return _validar_minuto_numerico(minutos)

    if isinstance(valor, bool):
        raise ValueError(
            "el valor booleano no representa una hora."
        )

    if isinstance(valor, (int, float)):
        numero = float(valor)

        if not isfinite(numero):
            raise ValueError(
                "la hora no es un número finito."
            )

        if 0.0 <= numero < 1.0:
            return _validar_minuto_numerico(
                numero * 24.0 * 60.0
            )

        return _validar_minuto_numerico(numero)

    texto = _texto(valor)

    if not texto:
        raise ValueError("la hora está vacía.")

    if texto.isdigit():
        return _validar_minuto_numerico(
            float(texto)
        )

    coincidencia = _RE_HORA.fullmatch(texto)

    if coincidencia is None:
        raise ValueError(
            "use una hora como 07:30 o 14:45."
        )

    return (
        int(coincidencia.group("hora")) * 60
        + int(coincidencia.group("minuto"))
    )


def formatear_minutos_hora(minutos: int) -> str:
    if minutos < 0 or minutos >= 24 * 60:
        raise ValueError(
            "Los minutos deben estar entre 0 y 1439."
        )

    return f"{minutos // 60:02d}:{minutos % 60:02d}"


def _parsear_fila(
    numero_fila: int,
    fila: dict[str, Any],
    *,
    turno: str,
) -> tuple[PedidoExcel | None, list[ErrorFilaExcel]]:
    errores: list[ErrorFilaExcel] = []

    pedido_id = _texto(
        fila.get("pedido_id")
    )

    if not pedido_id:
        errores.append(
            ErrorFilaExcel(
                numero_fila,
                "pedido_id",
                "es obligatorio.",
            )
        )

    unidades = _parsear_entero_positivo(
        fila.get("unidades"),
        numero_fila,
        "unidades",
        errores,
    )

    latitud = _parsear_decimal(
        fila.get("latitud"),
        numero_fila,
        "latitud",
        errores,
    )

    longitud = _parsear_decimal(
        fila.get("longitud"),
        numero_fila,
        "longitud",
        errores,
    )

    if (
        latitud is not None
        and not -90.0 <= latitud <= 90.0
    ):
        errores.append(
            ErrorFilaExcel(
                numero_fila,
                "latitud",
                "debe estar entre -90 y 90.",
            )
        )

    if (
        longitud is not None
        and not -180.0 <= longitud <= 180.0
    ):
        errores.append(
            ErrorFilaExcel(
                numero_fila,
                "longitud",
                "debe estar entre -180 y 180.",
            )
        )

    requiere_volcador = _parsear_booleano(
        fila.get("requiere_volcador"),
        numero_fila,
        "requiere_volcador",
        errores,
        default=False,
    )

    tiene_ventana = _parsear_booleano(
        fila.get("tiene_ventana"),
        numero_fila,
        "tiene_ventana",
        errores,
        default=False,
    )

    hora_inicio, hora_fin = HORARIOS_TURNO[turno]
    hora_desde = hora_inicio
    hora_hasta = hora_fin

    if tiene_ventana:
        hora_desde = _parsear_hora_campo(
            fila.get("hora_desde"),
            numero_fila,
            "hora_desde",
            errores,
        )

        hora_hasta = _parsear_hora_campo(
            fila.get("hora_hasta"),
            numero_fila,
            "hora_hasta",
            errores,
        )

        if (
            hora_desde is not None
            and hora_hasta is not None
        ):
            if hora_desde >= hora_hasta:
                errores.append(
                    ErrorFilaExcel(
                        numero_fila,
                        "ventana",
                        "hora_desde debe ser anterior "
                        "a hora_hasta.",
                    )
                )

            if (
                hora_desde < hora_inicio
                or hora_hasta > hora_fin
            ):
                errores.append(
                    ErrorFilaExcel(
                        numero_fila,
                        "ventana",
                        (
                            "debe quedar completamente "
                            f"dentro del turno {turno} "
                            f"({formatear_minutos_hora(hora_inicio)}"
                            "–"
                            f"{formatear_minutos_hora(hora_fin)})."
                        ),
                    )
                )

    if errores:
        return None, errores

    assert unidades is not None
    assert latitud is not None
    assert longitud is not None
    assert requiere_volcador is not None
    assert tiene_ventana is not None
    assert hora_desde is not None
    assert hora_hasta is not None

    return (
        PedidoExcel(
            fila_excel=numero_fila,
            pedido_id=pedido_id,
            cliente=_texto(fila.get("cliente")),
            direccion=_texto(
                fila.get("direccion")
            ),
            barrio=_texto(fila.get("barrio")),
            unidades=unidades,
            latitud=latitud,
            longitud=longitud,
            requiere_volcador=requiere_volcador,
            tiene_ventana=tiene_ventana,
            hora_desde_min=hora_desde,
            hora_hasta_min=hora_hasta,
            observaciones=_texto(
                fila.get("observaciones")
            ),
        ),
        [],
    )


def _seleccionar_hoja(
    libro: Any,
    hoja_solicitada: str | None,
) -> Any:
    if hoja_solicitada:
        if hoja_solicitada not in libro.sheetnames:
            raise ValueError(
                "No existe la hoja solicitada: "
                f"{hoja_solicitada}."
            )
        return libro[hoja_solicitada]

    if HOJA_PEDIDOS_PREFERIDA in libro.sheetnames:
        return libro[HOJA_PEDIDOS_PREFERIDA]

    return libro.active


def _detectar_encabezado(
    filas: Iterable[tuple[Any, ...]],
) -> tuple[int, dict[str, int]]:
    mejor: tuple[int, dict[str, int]] | None = None

    for numero_fila, valores in enumerate(
        filas,
        start=1,
    ):
        columnas: dict[str, int] = {}

        for indice, valor in enumerate(valores):
            normalizado = _normalizar_columna(valor)

            if not normalizado:
                continue

            for nombre, aliases in _ALIAS_COLUMNAS.items():
                if normalizado in aliases:
                    columnas.setdefault(nombre, indice)
                    break

        if (
            "pedido_id" in columnas
            and len(columnas) >= 6
        ):
            mejor = (numero_fila, columnas)
            break

    if mejor is None:
        raise ErrorImportacionExcel(
            [
                ErrorFilaExcel(
                    fila=1,
                    campo="encabezado",
                    mensaje=(
                        "no se encontró una fila de "
                        "encabezados válida dentro de las "
                        "primeras 20 filas."
                    ),
                )
            ]
        )

    return mejor


def _normalizar_columna(valor: Any) -> str:
    texto = unicodedata.normalize(
        "NFKD",
        _texto(valor).lower(),
    )

    sin_acentos = "".join(
        caracter
        for caracter in texto
        if not unicodedata.combining(caracter)
    )

    return "".join(
        caracter
        for caracter in sin_acentos
        if caracter.isalnum()
    )


def _normalizar_turno(turno: str) -> str:
    valor = _texto(turno).upper()

    if valor == "MAÑANA":
        valor = "MANANA"

    if valor not in HORARIOS_TURNO:
        raise ValueError(
            "Turno no soportado: "
            f"{turno}. Use MANANA o TARDE."
        )

    return valor


def _texto(valor: Any) -> str:
    if valor is None:
        return ""

    return str(valor).strip()


def _fila_vacia(valores: Iterable[Any]) -> bool:
    return all(not _texto(valor) for valor in valores)


def _parsear_entero_positivo(
    valor: Any,
    fila: int,
    campo: str,
    errores: list[ErrorFilaExcel],
) -> int | None:
    try:
        if isinstance(valor, bool):
            raise ValueError

        numero = float(
            _texto(valor).replace(",", ".")
        )

        if not isfinite(numero):
            raise ValueError

        entero = int(numero)

        if numero != entero or entero <= 0:
            raise ValueError

        return entero

    except (TypeError, ValueError):
        errores.append(
            ErrorFilaExcel(
                fila,
                campo,
                "debe ser un entero mayor que cero.",
            )
        )
        return None


def _parsear_decimal(
    valor: Any,
    fila: int,
    campo: str,
    errores: list[ErrorFilaExcel],
) -> float | None:
    try:
        if isinstance(valor, bool):
            raise ValueError

        numero = float(
            _texto(valor).replace(",", ".")
        )

        if not isfinite(numero):
            raise ValueError

        return numero

    except (TypeError, ValueError):
        errores.append(
            ErrorFilaExcel(
                fila,
                campo,
                "debe ser un número válido.",
            )
        )
        return None


def _parsear_booleano(
    valor: Any,
    fila: int,
    campo: str,
    errores: list[ErrorFilaExcel],
    *,
    default: bool,
) -> bool | None:
    if valor is None or _texto(valor) == "":
        return default

    if isinstance(valor, bool):
        return valor

    texto = _normalizar_columna(valor)

    if texto in {"si", "s", "true", "verdadero", "1"}:
        return True

    if texto in {"no", "n", "false", "falso", "0"}:
        return False

    errores.append(
        ErrorFilaExcel(
            fila,
            campo,
            "use SI o NO.",
        )
    )
    return None


def _parsear_hora_campo(
    valor: Any,
    fila: int,
    campo: str,
    errores: list[ErrorFilaExcel],
) -> int | None:
    try:
        return parsear_hora_a_minutos(valor)
    except ValueError as ex:
        errores.append(
            ErrorFilaExcel(
                fila,
                campo,
                str(ex),
            )
        )
        return None


def _validar_minuto_numerico(
    minutos: float,
) -> int:
    redondeado = int(round(minutos))

    if abs(minutos - redondeado) > 1e-6:
        raise ValueError(
            "la hora debe resolverse a minutos enteros."
        )

    if redondeado < 0 or redondeado >= 24 * 60:
        raise ValueError(
            "la hora debe estar entre 00:00 y 23:59."
        )

    return redondeado
